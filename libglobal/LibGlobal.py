from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select
import openpyxl


class Base:
    # WebDriver Methods
    def browser_launch(self):
        self.driver=webdriver.Chrome(executable_path=r"C:\Users\Gobinath\PycharmProjects\PythonLearning\Drivers\chromedriver.exe")
        return self.driver

    def load_url(self, url):
        self.driver.get(url)

    def window_maximize(self):
        self.driver.maximize_window()

    def win_min(self):
        self.driver.minimize_window()

    def win_quit(self):
        self.driver.quit()

    def win_close(self):
        self.driver.close()

    def win_refresh(self):
        self.driver.refresh()

    def win_back(self):
        self.driver.back()

    def win_forward(self):
        self.driver.forward()

    def screen_shot(self, name):
        self.driver.save_screenshot(name)

    def jscript(self,elementjs, data):
        self.driver.execute_script("arguments[0].setAttribute('value', '{}')".format(data), elementjs)


    # WebDriver Parameters:

    def page_url(self):
        url = self.driver.current_url
        return url

    def page_title(self):
        t=self.driver.title
        return t



# WebElement Methods
    def send_input(self, elementinput, data):
        elementinput.send_keys(data)

    def txt(self, elementatt):
        att = elementatt.get_attribute('value')
        return att

    def btn_click(self, elementclick):
        elementclick.click()

    def btn_submit(self, elementsubmit):
        elementsubmit.submit()

    def clr_val(self, elementclr):
        elementclr.clear()

    def screen_shot2(self, webeleloc):
        self.driver.save_screenshot("webeleloc")

    def selected(self, elementselected):
        s=elementselected.is_selected()
        return s

    def enabled(self, elementenabled):
        e=elementenabled.is_enabled()
        return e

    def displayed(self, elementdisplayed):
        d=elementdisplayed.is_displayed()
        return d

# WebElement Parameter:

    def get_text(self, elementtxt):
        t=elementtxt.text
        return t


# ActionChains Methods

    def mouse_move(self, elementmove):
        ac=ActionChains(self.driver)
        ac.move_to_element(elementmove).perform()

    def drag_drop(self, src, des):
        ac = ActionChains(self.driver)
        ac.drag_and_drop(src, des).perform()

    def dbl_click(self, elementdoubleclick):
        ac=ActionChains(self.driver)
        ac.double_click(elementdoubleclick).perform()

    def right_click(self, elementrightclick):
        ac=ActionChains(self.driver)
        ac.context_click(elementrightclick).perform()

    # Select Methods

    def index_select(self, elementsel, index):
        s = Select(elementsel)
        s.select_by_index(index)

    def value_select(self, elementval, value):
        s = Select(elementval)
        s.select_by_value(value)

    def text_select(self, elementtxt, txt):
        s = Select(elementtxt)
        s.select_by_visible_text(txt)

    def index_deselect(self, elementdesel, index):
        s= Select(elementdesel)
        s.deselect_by_index(index)

    def value_deselect(self, elementdeselval, value):
        s=Select(elementdeselval)
        s.deselect_by_value(value)

    def text_deselect(self, elementtxtdesel, txt):
        s=Select(elementtxtdesel)
        s.deselect_by_visible_text(txt)

    def all_deselect(self, elementdeselectall):
        s=Select(elementdeselectall)
        s.deselect_all()

    def all_opt(self, elementall):
        s=Select(elementall)
        op=s.options
        return op

    def first_selected(self, elementall):
        s=Select(elementall)
        f=s.first_selected_option.text
        return f

    def all_selected(self, elementall):
        s=Select(elementall)
        all=s.all_selected_options
        for i in all:
            return i.text


    def multiple(self, element):
        s= Select(element)
        multi = s.is_multiple
        return multi

    # Alerts Methods
    def simple_alert(self):
        al=self.driver.switch_to.alert
        al.accept()

    def confirm_alert_acpt(self):
        al=self.driver.switch_to.alert
        al.accept()

    def confirm_alert_dismiss(self):
        al=self.driver.switch_to.alert
        al.dismiss()

    def prompt_alert_acpt(self, data):
        al=self.driver.switch_to.alert
        al.send_keys(data)
        al.accept()

    def prompt_alert_dismiss(self):
        al=self.driver.switch_to.alert
        al.dismiss()

    # Alert Paramter
    def prompt_alert_txt(self):
        al=self.driver.switch_to.alert
        t=al.text
        return t

    # Data Driven

    def get_data_from_excel(self, row, col, worksheet):
        excel_loc = r"C:\Users\Gobinath\OneDrive\Desktop\Test.xlsx"
        w = openpyxl.load_workbook(excel_loc)
        sheet = w.get_sheet_by_name(worksheet)
        c = sheet.cell(row, col)
        data = c.value

        return data

    def get_max_row(self):
        excel_loc = r"C:\Users\Gobinath\OneDrive\Desktop\Test.xlsx"
        w = openpyxl.load_workbook(excel_loc)
        sheet = w.active
        all_row = sheet.max_row
        return all_row

    def get_max_col(self):
        excel_loc = r"C:\Users\Gobinath\OneDrive\Desktop\Test.xlsx"
        w = openpyxl.load_workbook(excel_loc)
        sheet = w.active
        all_col = sheet.max_column
        return all_col















