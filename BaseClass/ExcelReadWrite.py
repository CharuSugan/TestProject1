import time

import keyboard
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

# excel_loc=r"C:\Users\Gobinath\OneDrive\Desktop\Python_Tasks\Data_Driven.xlsx"
# print("----------Excel Read-----------")
# w=openpyxl.load_workbook(excel_loc)
# sheet=w["Data1"]
# c=sheet.cell(1,1)
# data=c.value
# print(data)
#
# print("----------Excel Write-----------")
# w1=openpyxl.Workbook()
# excel_create="Sample.xlsx"
# sheet=w1.create_sheet("Sample1", 0)
# c=sheet.cell(1,1)
# c.value = "Charu"
# w1.save(excel_create)

driver=webdriver.Chrome(executable_path=r"C:\Users\Gobinath\PycharmProjects\PythonLearning\Drivers\chromedriver.exe")
driver.implicitly_wait(20)

excel_loc=r"C:\Users\Gobinath\OneDrive\Desktop\Python_Tasks\ExcelReadWrite.xlsx"
w=openpyxl.load_workbook(excel_loc)

# Demo Testing Lanugage Write
driver.get("http://demo.automationtesting.in/Register.html")
btn_lang=driver.find_element(By.ID, "msdd")
btn_lang.click()
lang_list=driver.find_elements(By.XPATH, "//a[@class='ui-corner-all']")
sheet=w["Demo"]

i=1
for val in lang_list:
    data=val.text
    print(data)
    c=sheet.cell(i, 1)
    c.value=data
    w.save(excel_loc)
    i=i+1

time.sleep(10)

# Amazon Search Result
driver.get("https://www.amazon.in/")

# w1=openpyxl.load_workbook(excel_loc)
search_box=driver.find_element(By.ID, "twotabsearchtextbox")
search_box.send_keys("iphone")
keyboard.press("Enter")
keyboard.release("Enter")
search_result=driver.find_elements(By.XPATH, "//span[@class='a-size-medium a-color-base a-text-normal']")
sheet=w["Amazon"]

j=1
for result in search_result:
    phone_names=result.text
    print(phone_names)
    c=sheet.cell(j, 1)
    c.value=phone_names
    w.save(excel_loc)
    j=j+1








